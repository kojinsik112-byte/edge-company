# -*- coding: utf-8 -*-
# gen_contract.py 의 DATA/VENDORS/YAKGWAN 를 그대로 엑셀로 (회장 직접 편집용)
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# --- gen_contract.py 에서 데이터만 가져오기 (HTML 쓰기 부작용 억제) ---
src = open("/tmp/claude-0/-home-user-edge-company/934a8e34-8e01-597a-96a6-1875710c39c8/scratchpad/gen_contract.py", encoding="utf-8").read()
src = re.sub(r'^open\(.*write\(html\).*$', '', src, flags=re.M)
src = re.sub(r'^print\("written"\).*$', '', src, flags=re.M)
ns = {}
exec(compile(src, "gen_contract", "exec"), ns)
DATA, VENDORS, YAKGWAN, pays, won = ns["DATA"], ns["VENDORS"], ns["YAKGWAN"], ns["pays"], ns["won"]

def clean(s):  # HTML 태그/엔티티 제거 → 순수 텍스트
    s = re.sub(r'<br\s*/?>', ' / ', s)
    s = re.sub(r'<[^>]+>', '', s)
    s = s.replace('&amp;', '&').replace('&ndash;', '-').replace('&nbsp;', ' ').replace('&rarr;', '→')
    return re.sub(r'\s+', ' ', s).strip()

NAVY = "15233D"; CREAM = "FFF7DF"; GREY = "EEF0F4"
thin = Side(style="thin", color="B3B8C2")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
cen = Alignment(horizontal="center", vertical="center", wrap_text=True)
left = Alignment(horizontal="left", vertical="center", wrap_text=True)
right = Alignment(horizontal="right", vertical="center")
hdrfont = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")
navyfont = Font(name="맑은 고딕", size=10, bold=True, color=NAVY)
redfont = Font(name="맑은 고딕", size=9, color="C0392B")
basefont = Font(name="맑은 고딕", size=10)

wb = Workbook()

# ================= 시트1: 계약서 (단가표) =================
ws = wb.active; ws.title = "옵션계약서"
cols = ["순번","품목","Type","제품명 / 사양","총금액","계약금(20%)","중도금(10%)","잔금(70%)","계약자날인","비고"]
widths = [6, 12, 12, 34, 13, 12, 12, 13, 11, 40]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.merge_cells("A1:J1")
t = ws["A1"]; t.value = "힐스테이트 물금센트럴 추가 옵션 계약서 · 제2조 (옵션품목 및 공급대금)  [단위: 원 / VAT 포함]"
t.font = Font(name="맑은 고딕", size=13, bold=True, color=NAVY); t.alignment = cen
ws.row_dimensions[1].height = 26

r = 2
for c, name in enumerate(cols, 1):
    cell = ws.cell(row=r, column=c, value=name)
    cell.font = hdrfont; cell.fill = PatternFill("solid", fgColor=NAVY); cell.alignment = cen; cell.border = border
ws.row_dimensions[r].height = 30

r = 3
for no, pname, remark, groups in DATA:
    start = r
    for prod, rows in groups:
        gstart = r
        for typ, amt in rows:
            ws.cell(row=r, column=3, value=clean(typ)).alignment = cen
            if amt == "선택":
                ws.cell(row=r, column=5, value="선택").alignment = cen
                for cc in (6,7,8): ws.cell(row=r, column=cc, value="-").alignment = cen
            else:
                g1,g2,g3 = [int(x.replace(',','')) for x in pays(amt)]
                ws.cell(row=r, column=5, value=amt).alignment = right
                ws.cell(row=r, column=6, value=g1).alignment = right
                ws.cell(row=r, column=7, value=g2).alignment = right
                ws.cell(row=r, column=8, value=g3).alignment = right
            ws.cell(row=r, column=9, value="(인)").alignment = cen
            for cc in range(1,11):
                cell = ws.cell(row=r, column=cc); cell.border = border
                if cell.font.name is None or cell.font.size is None: cell.font = basefont
                if cc in (5,6,7,8): cell.number_format = '#,##0'
            ws.row_dimensions[r].height = 34
            r += 1
        ws.merge_cells(start_row=gstart, start_column=4, end_row=r-1, end_column=4)
        pc = ws.cell(row=gstart, column=4, value=clean(prod)); pc.alignment = left; pc.font = basefont
    # 순번/품목/비고 병합
    ws.merge_cells(start_row=start, start_column=1, end_row=r-1, end_column=1)
    ws.cell(row=start, column=1, value=no).font = navyfont
    ws.merge_cells(start_row=start, start_column=2, end_row=r-1, end_column=2)
    bc = ws.cell(row=start, column=2, value=clean(pname)); bc.font = navyfont; bc.fill = PatternFill("solid", fgColor=CREAM); bc.alignment = cen
    ws.merge_cells(start_row=start, start_column=10, end_row=r-1, end_column=10)
    rc = ws.cell(row=start, column=10, value=clean(remark)); rc.font = redfont; rc.alignment = left

# ================= 시트2: 업체 입금계좌 =================
ws2 = wb.create_sheet("업체계좌")
for i, w in enumerate([12, 26, 50, 22, 10], 1):
    ws2.column_dimensions[get_column_letter(i)].width = w
ws2.merge_cells("A1:E1")
h = ws2["A1"]; h.value = "납부방법 · 품목별 대표 업체 입금계좌"
h.font = Font(name="맑은 고딕", size=13, bold=True, color=NAVY); h.alignment = cen
for c, name in enumerate(["금융기관","예금주(공급업체)","담당 품목","계좌번호","도장"], 1):
    cell = ws2.cell(row=2, column=c, value=name); cell.font = hdrfont; cell.fill = PatternFill("solid", fgColor=NAVY); cell.alignment = cen; cell.border = border
ws2.row_dimensions[2].height = 26
r = 3
for item, ven, bk, acct in VENDORS:
    ws2.cell(row=r, column=1, value=clean(bk)).alignment = cen
    ws2.cell(row=r, column=2, value=clean(ven)).alignment = cen
    ws2.cell(row=r, column=3, value=clean(item)).alignment = left
    ac = ws2.cell(row=r, column=4, value=acct); ac.alignment = cen; ac.font = Font(name="맑은 고딕", size=12, bold=True, color=NAVY)
    ws2.cell(row=r, column=5, value="(인)").alignment = cen
    for cc in range(1,6): ws2.cell(row=r, column=cc).border = border
    ws2.row_dimensions[r].height = 30
    r += 1
ws2.cell(row=r+1, column=1, value="※ 입금 시 반드시 「동·호수 + 계약자 성명」 기재 (예: 101동 1001호 홍길동 → 1011001홍길동) · 각 대표 업체 계좌로 직접 납부 · 엣지컴퍼니 이행·품질·A/S 보증 · 계약 취소는 계약일로부터 15일(2주) 이내만 가능").font = redfont

# ================= 시트3: 약관 =================
ws3 = wb.create_sheet("약관")
ws3.column_dimensions["A"].width = 110
ws3["A1"] = "【 약관 】 힐스테이트 물금센트럴 추가 옵션 계약서"; ws3["A1"].font = Font(name="맑은 고딕", size=13, bold=True, color=NAVY)
r = 3
for title, items in YAKGWAN:
    ws3.cell(row=r, column=1, value=title).font = navyfont; r += 1
    for i, it in enumerate(items, 1):
        c = ws3.cell(row=r, column=1, value=f"  {i}. {clean(it)}"); c.font = basefont; c.alignment = Alignment(wrap_text=True, vertical="top"); r += 1
    r += 1

out = "/home/user/edge-company/business-ops/movein/양산현장_서류/힐스테이트물금센트럴_추가옵션계약서.xlsx"
wb.save(out)
print("saved", out)
